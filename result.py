import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook  # Correct import for Workbook
from openpyxl.drawing.image import Image  # Correct import for inserting images into Excel
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np


def convert_marks_to_grade(marks):
    try:
        marks = float(marks)  # Convert to number
    except ValueError:
        return marks, 0  # If non-numeric (like '-', 'A', 'B'), return as is

    if marks >= 90:
        return "O", 10
    elif marks >= 80:
        return "A+", 9
    elif marks >= 70:
        return "A", 8
    elif marks >= 60:
        return "B+", 7
    elif marks >= 55:
        return "B", 6
    elif marks >= 50:
        return "C", 5
    else:
        return "F", 0

def calculate_sgpa(grades, credits):
    total_points = sum(grade * credit for grade, credit in zip(grades, credits))
    total_credits = sum(credits)
    return round(total_points / total_credits, 2) if total_credits > 0 else 0

def determine_class(sgpa):
    if sgpa >= 7:
        return "FCD"
    elif sgpa >= 6:
        return "FC"
    elif sgpa >= 5:
        return "SC"
    else:
        return "Fail"

def process_results():
    df = pd.read_csv("result.csv")
    subjects = df.columns[3:]  # Extract subject columns starting from the 4th column

    subject_credits = {
        "P22MCA21": 4, "P22MCA22": 4, "P23MCA23": 4, "P22MCA24": 4,
        "P22MCA251": 3, "P22MCA261": 3, "P22MCAL27": 1, "P22MCAL28": 1,
        "P22MCA255": 3, "P22MCA265": 3
    }

    # Replace `0` or missing values with `-` for all subjects
    for subj in subjects:
        if subj in df.columns:
            df[subj] = df[subj].replace(0, "-").fillna("-")

    # Ensure subject columns are processed as strings
    for subj in subjects:
        df[subj] = df[subj].astype(str)

    sgpas, classes, results = [], [], []

    for index, row in df.iterrows():
        grades, points, credits = [], [], []

        for subj in subjects:
            mark = row[subj]
            if mark == '-' or not str(mark).replace('.', '', 1).isdigit():
                grades.append('NA')  # Skip invalid or missing values
                points.append(0)
                credits.append(0)
            else:
                grade, point = convert_marks_to_grade(float(mark))
                grades.append(grade)
                points.append(point)
                credits.append(subject_credits.get(subj, 0))  # Assign correct credit

        sgpa = calculate_sgpa(points, credits)
        student_class = determine_class(sgpa)

        # Updated condition for determining "Pass" or "Fail"
        if "F" in grades and sgpa < 5.0:
            result_status = "Fail"
        else:
            result_status = "Pass"

        sgpas.append(sgpa)
        classes.append(student_class)
        results.append(result_status)

        for i, subj in enumerate(subjects):
             df.at[index, subj] = grades[i]

    df["SGPA"], df["CLASS"], df["RESULT"] = sgpas, classes, results
    df.to_csv("processed_result.csv", index=False)  # Save the updated data

    return df

def generate_result_sheet(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Result Sheet"

    # Set column widths for B and C
    ws.column_dimensions['B'].width = 20  # Adjust width as needed
    ws.column_dimensions['C'].width = 25  # Adjust width as needed
    # Set column widths for B and C
    ws.column_dimensions['A'].width = 20  # Adjust width as needed
    ws.column_dimensions['E'].width = 25  # Adjust width as needed

    # Add the header with college name and address
    ws.merge_cells("A1:N1")
    ws["A1"] = "PES COLLEGE OF ENGINEERING,MANDYA-571 401"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = "(An Autonomous Institution Affiliated to VTU, Belagavi.)"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:N3")
    ws["A3"] = "Provisional Results of SECOND Semester MCA"
    ws["A3"].font = Font(bold=True, size=15)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:N4")
    ws["A4"] = "Semester End Examination for Academic year 2023-24"
    ws["A4"].font = Font(bold=True, size=16)
    ws["A4"].alignment = Alignment(horizontal="center")

    # Leave a blank line
    ws.merge_cells("A5:N5")

    # Add Course/Branch, Year/Sem, Batch, and Date
    ws["A6"] = "Course/Branch:"
    ws["A6"].font = Font(bold=True, size=13)
    ws["B6"] = "MCA"
    ws["B6"].font = Font(bold=True, size=13)
    ws["D6"] = "Year/Sem:"
    ws["D6"].font = Font(bold=True, size=13)
    ws["E6"] = "1st Year / II Sem"
    ws["E6"].font = Font(bold=True, size=13)
    ws["G6"] = "Batch:"
    ws["G6"].font = Font(bold=True, size=13)
    ws["H6"] = "2023-2025"
    ws["H6"].font = Font(bold=True, size=13)
    ws["J6"] = "Date:"
    ws["J6"].font = Font(bold=True, size=13)
    ws["K6"] = "11-02-2025"
    ws["K6"].font = Font(bold=True, size=13)

    # Add logos
    left_logo = Image("logo-left.png")
    right_logo = Image("logo-right.png")
    # Resize logos
    left_logo.width, left_logo.height = 100, 100
    right_logo.width, right_logo.height = 100, 100
    ws.add_image(left_logo, "A2")
    ws.add_image(right_logo, "N2")

    # Add column headers
    for col_num, column_title in enumerate(df.columns, start=1):
        ws.cell(row=8, column=col_num, value=column_title).font = Font(bold=True)

    # Add student data
    for row_num, row_data in enumerate(df.itertuples(index=False), start=9):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save("result_sheet.xlsx")




def calculate_cgpa(df):
    """Calculate CGPA for students based on SGPA values."""
    sgpa_columns = [col for col in df.columns if 'SGPA' in col]
    df['CGPA'] = df[sgpa_columns].mean(axis=1).round(2)
    return df




import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from io import BytesIO

def generate_result_analysis(input_file):
    # Load the processed results CSV file
    df = pd.read_csv(input_file)

    # Normalize column names
    df.columns = df.columns.str.strip().str.upper()

    # Required columns
    required_columns = ['NAME', 'SGPA', 'P22MCA21', 'P22MCA22', 'P22MCA23', 'P22MCA24', 'P22MCAL27', 'P22MCAL28']
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column: {col}")

    # Convert subject marks to numeric
    subject_cols = ['P22MCA21', 'P22MCA22', 'P22MCA23', 'P22MCA24', 'P22MCAL27', 'P22MCAL28']
    df[subject_cols] = df[subject_cols].apply(pd.to_numeric, errors='coerce')

    # Sort by SGPA (Top 5 students)
    top_students = df[['NAME', 'SGPA']].sort_values(by='SGPA', ascending=False).head(5)

    # Pass Percentage
    total_students = len(df)
    passed_students = df[df['SGPA'] >= 5.0]
    num_passed = len(passed_students)
    pass_percentage = (num_passed / total_students) * 100

    # Division Classification
    honors_count = len(df[df['SGPA'] >= 8.5])
    first_division_count = len(df[(df['SGPA'] >= 7.0) & (df['SGPA'] < 8.5)])
    second_division_count = len(df[(df['SGPA'] >= 5.0) & (df['SGPA'] < 7.0)])

    # Pass percentage per subject
    pass_percentages = [(df[subject] >= 40).mean() * 100 for subject in subject_cols]

    # Average Marks per Subject
    average_marks = [df[subject].mean() for subject in subject_cols]

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Result Analysis"

    # Add College Name, Address, and Logos
    title_font = Font(size=14, bold=True)
    subtitle_font = Font(size=12, bold=True)
    ws.merge_cells("A1:E1")
    ws["A1"] = "PES COLLEGE OF ENGINEERING, MANDYA"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = "(An Autonomous Institution Affiliated to VTU, Belagavi.)"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:E3")
    ws["A3"] = "DEPARTMENT OF MASTER OF COMPUTER APPLICATION"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:E4")
    ws["A4"] = "MCA / Computer Science - 1st Year / 2nd Sem 2023-2024"
    ws["A4"].font = subtitle_font
    ws["A4"].alignment = Alignment(horizontal="center")

    ws.append([])  # Empty row for spacing

    # Insert logos
    left_logo = Image("logo-left.png")  
    right_logo = Image("logo-right.png")  
    left_logo.width, left_logo.height = 110, 120
    right_logo.width, right_logo.height = 110, 120
    ws.add_image(left_logo, "A1")
    ws.add_image(right_logo, "E1")

    # Add Pass Statistics
    ws.append(["Pass Statistics"])
    ws["A6"].font = Font(bold=True, size=18)
    stats = [
        ["Pass Percentage", f"{pass_percentage:.2f}%"],
        ["Total Students Appeared", total_students],
        ["No. of Students Passed", num_passed],
        ["No. of Students Passed with Honors", honors_count],
        ["No. of Students Passed in 1st Division", first_division_count],
        ["No. of Students Passed in 2nd Division", second_division_count],
    ]
    for row in stats:
        ws.append(row)

    # Make key statistics bold
    for cell in ws["A7:A12"]:
        cell[0].font = Font(bold=True)

    ws.append([])

    # Add Top Five Students
    ws.append(["Top Five Students"])
    ws["A14"].font = Font(bold=True, size=18)
    ws.append(["Rank", "Name", "SGPA"])
    for cell in ws["A15:C15"]:
        cell[0].font = Font(bold=True)
    for i, row in enumerate(top_students.itertuples(), start=1):
        ws.append([i, row.NAME, row.SGPA])

    ws.append([])

    # Add Subject-wise Pass Percentage
    ws.append(["Pass Percentage per Subject"])
    ws["A21"].font = Font(bold=True, size=18)
    ws.append(["Subject", "Pass Percentage"])
    for cell in ws["A22:B22"]:
        cell[0].font = Font(bold=True, size=18)

    # Define professional color palette for subjects
    subject_colors = ["#4E79A7", "#F28E2B", "#E15759", "#76B7B2", "#59A14F", "#EDC949"]

    # Add pass percentages for each subject
    for i, subject in enumerate(subject_cols):
        ws.append([subject, f"{pass_percentages[i]:.2f}%"])
        ws[f"A{ws.max_row}"].font = Font(size=12)
        ws[f"B{ws.max_row}"].font = Font(size=12)

    ws.append([])

    # Add Subject-wise Average Marks
    ws.append(["Average Marks per Subject"])
    ws["A28"].font = Font(bold=True, size=18)
    ws.append(["Subject", "Average Marks"])
    for cell in ws["A29:B29"]:
        cell[0].font = Font(bold=True, size=18)

    for i, subject in enumerate(subject_cols):
        ws.append([subject, f"{average_marks[i]:.2f}"])

    ws.append([])

    # Generate Graphs with different colors for each subject
    def save_plot(data, title, colors, ylabel):
        fig, ax = plt.subplots()
        bars = ax.bar(subject_cols, data, color=colors)
        
        # Label each bar with its percentage/marks
        for bar, value in zip(bars, data):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1, f"{value:.2f}",
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

        ax.set_title(title, fontsize=14, fontweight="bold")
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_xticklabels(subject_cols, rotation=45, ha="right", fontsize=10)
        ax.grid(axis='y', linestyle='--', alpha=0.7)

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format="png", bbox_inches="tight")
        plt.close()
        return Image(img_buffer)

    img1 = save_plot(pass_percentages, "Pass Percentage of Students", subject_colors, "Pass Percentage (%)")
    img2 = save_plot(average_marks, "Average Marks per Subject", subject_colors, "Marks")

    # Insert Graphs at the bottom
    ws.add_image(img1, "A41")
    ws.add_image(img2, "A67")

    # Adjust Column Widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Save Excel File
    excel_filename = "result_analysis.xlsx"
    wb.save(excel_filename)
    print(f"✅ Result Analysis saved as {excel_filename}")

# Run the function
generate_result_analysis("result_analysis.csv")






def main():
    df = process_results()
    generate_result_sheet(df)
    generate_result_analysis("result_analysis.csv")

if __name__ == "__main__":
    main()
